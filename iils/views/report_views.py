from time import timezone

from django.shortcuts import render, get_object_or_404
from openpyxl import Workbook
from io import BytesIO
import pandas as pd
from datetime import datetime

now = datetime.now()

# Create your views here.
wb = Workbook()  # 엑셀 생성
ws = wb.active	# 엑셀 활성화
excelfile = BytesIO() #바이트 배열 생성

ws.title = "제품분류표준" # 엑셀 a1 열 이름 정하기
#ws['A1']= '제품명'
# 시트에 추가할 칼럼의 목록을 리스트형으로 'column'이라는 변수에 지정
column = ['번호', 'GPCK', 'HSK']

# column리스트 목록을 시트 첫 행에 입력
ws.append(column)

# 시트에 추가할 데이터를 리스트형으로 'row'라는 변수에 지정
row = [1, '7000000', '10010001']

# append로 row의 목록을 column 아래 행에 입력
ws.append(row)

# '인증정보', '유해정보', '리콜정보', '수출입정보' 시트 추가
wb.create_sheet('인증정보')
wb.create_sheet('유해정보')
wb.create_sheet('수출입정보')

# 파일 명에 맞는 정규 표현식 패턴
filename = now.strftime('%Y-%m-%d')

wb.save('제품분류표준 '+filename+".xlsx")

wb.close()