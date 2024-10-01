from time import timezone

from django.shortcuts import render, get_object_or_404
from openpyxl import Workbook
from openpyxl import load_workbook
from io import BytesIO
import pandas as pd

import os
import re

# '관세청_HSK별 신성질별_성질별 분류_20240101.xlsx' 엑셀 파일 불러오기
wb = load_workbook(filename='관세청_HSK별_신성질별_성질별_분류_20240101.xlsx', data_only=True)

# 첫 번째 시트 불러오기
ws = wb.active
ws_title=ws.title
print(ws_title)

# Create your views here.
new_wb = Workbook()  # 엑셀 생성
new_ws = new_wb.active	# 엑셀 활성화
excelfile = BytesIO() #바이트 배열 생성


index_row = [cell.value for cell in list(ws.rows)[1]]  # 원본 파일의 두 번째 행(인덱스)
    new_rows = list(ws.rows)[2:] # 처음 2개 행을 제외하고 나머지 행을 리스트로 변환

    # 데이터 분류 및 저장
    for row in new_rows:
        if row[5].value == '신규':  # F열이 '신규'인 경우
            product = row[1].value  # B열의 제품 이름 가져와 product에 할당하기

            # 해당 제품 시트가 없으면 새로 생성
            if product not in new_wb.sheetnames:
                new_wb.create_sheet(title=product)
                product_ws = new_wb[product]
                product_ws.append(index_row)  # 인덱스 행을 새 시트의 첫 번째 행에 추가

            # 해당 시트를 선택
            product_ws = new_wb[product]

            # 데이터 추가
            product_ws.append([cell.value for cell in row])

# 임시로 만든 시트 삭제
del new_wb['Sheet']

# 파일 명에 맞는 정규 표현식 패턴
pattern = re.compile(r'\d{4}-\d{2}-\d{2} \d{4}\.xlsx')

# 새로운 엑셀 파일 저장
new_wb.save(filename='pattern')


